import openpyxl
import os
from functools import lru_cache
from dotenv import load_dotenv

load_dotenv()

class Config:
    EXCEL_PATH = os.path.join('predefined', 'user_predefined_inputs.xlsx')
    SHEET_NAME = 'Authentication'

    # local caching (for heavy objects)
    TIMEOUT =    os.environ.get('CACHE_TIMEOUT', 60) # 60 min
    LOCAL_CACHE = {
        'CACHE_TYPE':               'simple',
        'CACHE_DEFAULT_TIMEOUT':    TIMEOUT * 60
    }

    @classmethod
    @lru_cache(maxsize=1)
    def _EXCEL_WB(cls):
        return openpyxl.load_workbook(cls.EXCEL_PATH, data_only=True)

    @classmethod
    def load_authentication(cls):
        '''
        Load NoSQL database authentication info
        '''
        auth: dict[str, str] = {}

        # From local excel
        # wb = openpyxl.load_workbook(Config.EXCEL_PATH, data_only=True)
        # sheet = wb[Config.SHEET_NAME]

        # for row in sheet.iter_rows(min_row=2, values_only=True):
        #     platform, attribute, value = row
        #     if not isinstance(attribute, str):
        #         raise TypeError('Authentication file format error')
        #     auth[attribute.strip()] = value

        # From .env
        conn_str = os.environ.get('CONNECTION_STR')
        user = os.environ.get('USERNAME')
        pw = os.environ.get('PASSWORD')

        # Optional: Build full URI if missing
        if (
            'URI' not in auth
            and conn_str is not None
            and user is not None
            and pw is not None
        ):
            
            # Assume DB name == app name
            if '/?' in conn_str and 'appName' in conn_str:
                base, query_str = conn_str.split('/?', 1)
                query_params = query_str.split('&')

                dbname = None
                for param in query_params:
                    if param.startswith('appName='):
                        dbname = param.split('=', 1)[1]
                
                # Rebuild full connection string
                conn_str_rev = f'{base}/{dbname}?{query_str}'
            else:
                conn_str_rev = conn_str

            uri = (conn_str_rev).replace('<db_username>', user).replace('<db_password>', pw)

            auth = {
                'CONNECTION_STR':   conn_str,
                'USERNAME':         user,
                'PASSWORD':         pw,
                'URI':              uri,
            }

        return auth
    
    @property
    def _MONGO_URI(self):
        auth = self.load_authentication()
        uri = auth.get('URI')
        if not uri:
            raise ValueError("Mongo URI could not be constructed. Check your .env or Excel auth config.")
        return uri
    

