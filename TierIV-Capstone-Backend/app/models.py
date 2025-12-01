from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator, ValidationInfo
from typing import List, Union, Any, ClassVar
from config import Config
from app.service.query.query_models import MapLocation
import os
import openpyxl
import math
from functools import lru_cache
import zlib
import json
from shapely.geometry.base import BaseGeometry
from shapely.geometry import shape, mapping
from shapely import wkt
from enum import Enum


class WebRequest(BaseModel):
    '''
    Initial frontend webrequest base model
    - Configurated with backend user predefined inputs excel file
    - Default settings are set in model as fallbacks if the parameters are not present in the passed request
    - Validations added to ensure integrity of the populated fields
    - Adjustable analysis-specific parameters (e.g., junction angle thresholds) for backend modifications
    '''


    EXCEL_PATH: ClassVar[str]      = Config.EXCEL_PATH
    SHEET_QUERY: ClassVar[str]     = 'Default_Query'
    SHEET_SETTINGS: ClassVar[str]  = 'Manual_Settings'
    SHEET_ODD: ClassVar[str]       = 'ODD'
    SHEET_CONFLICT: ClassVar[str]   = 'Conflict_Classifier'


    @lru_cache(maxsize=None)
    def _EXCEL_WB():
        return openpyxl.load_workbook(WebRequest.EXCEL_PATH, data_only=True)


    # Operation
    overwrite: bool = Field(True, description='If `True`, overwrite existing results in database')
    default_query: bool = Field(True, description='If `True`, load query parameters from Excel')
    default_settings:   bool = Field(True, description='If `True`, use parameters defined in model; otherwise, load settings from Excel')
    default_conflict_classifier: bool = Field(True, description='If `True`, use parameters defined in model; otherwise, load settings from Excel')
    odd_all: bool = Field(True, description='If True, use parameters defined in model; otherwise, load from Excel')


    # Query
    input_type: str = Field(..., description='One of BBOX, POINT, ADDRESS, PLACE')
    input: Union[tuple, str] = Field(..., description='Coordinates or free-form address/place string')
    dist: float = Field(10000.0, ge=0.0, description='Search radius (for POINT/ADDRESS)')


    @model_validator(mode='before')
    def apply_default_query(cls, data) -> dict[str,Any]:
        # only act if asked
        if not data.get('default_query', True):
            return data
    
        wb    = WebRequest._EXCEL_WB()
        sheet = wb[cls.SHEET_QUERY]

        # read all rows into a dict {attribute_name: cell_value}
        row_map = {}
        for attr, val in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            if isinstance(attr, str):
                key = attr.strip().upper()
                row_map[key] = val

        # fill common fields
        # input_type is already set by user or default
        itype = data.get('input_type', row_map.get('INPUT_TYPE'))
        if not itype:
            raise ValueError('No INPUT_TYPE found in request or Excel')
        data['input_type'] = itype

        # choose which Excel cells to use for `input` and `dist`
        if itype == MapLocation.BBOX.name:
            try:
                bbox = (
                    float(row_map['MIN_LON (WEST)']),
                    float(row_map['MIN_LAT (SOUTH)']),
                    float(row_map['MAX_LON (EAST)']),
                    float(row_map['MAX_LAT (NORTH)'])
                )
            except KeyError:
                raise ValueError('Missing `BBOX` data columns in Excel')
            data.setdefault('input', bbox)

        elif itype == MapLocation.POINT.name:
            # CENTRAL_LAT, CENTRAL_LON
            lat = row_map.get('CENTRAL_LAT')
            lon = row_map.get('CENTRAL_LON')
            if lat is None or lon is None:
                raise ValueError('Missing `POINT` data columns in Excel')
            data.setdefault('input', [float(lat), float(lon)])
            # distance
            dist = row_map.get('DISTANCE (METERS)')
            if dist is not None:
                data.setdefault('dist', float(dist))

        elif itype == MapLocation.ADDRESS.name:
            addr = row_map.get('ADDRESS')
            if not addr:
                raise ValueError('Missing `ADDRESS` data in Excel')
            data.setdefault('input', addr)

            # allow distance too
            dist = row_map.get('DISTANCE (METERS)')
            if dist is not None:
                data.setdefault('dist', float(dist))

        elif itype == MapLocation.PLACE.name:
            place = row_map.get('PLACE')
            if not place:
                raise ValueError('Missing `PLACE` data in Excel')
            data.setdefault('input', place)

        return data
    

    @field_validator('input_type')
    def check_input_type(cls, v):
        if v not in ['BBOX', 'POINT', 'ADDRESS', 'PLACE']:
            raise ValueError(f'Query input type not compliant, must be one of [`BBOX`, `POINT`, `ADDRESS`, `PLACE`]')
        return v
        

    @field_validator('input')
    def check_input_shape(cls, v, info: ValidationInfo):
        t = info.data.get('input_type')
        if t == MapLocation.BBOX.name:
            if not(
                isinstance(v, tuple)
                and len(v) == 4
                and all(isinstance(x, (int, float)) for x in v)
                and v[0] < v[2]
                and v[1] < v[3]
            ):
                raise ValueError('BBOX inputs must be (west, south, east, north)')
            
            min_lon, min_lat, max_lon, max_lat = v
            mid_lat = (min_lat + max_lat) / 2.0

            # approximate meters per degree
            m_per_deg_lat = 110574.0
            m_per_deg_lon = 111320.0 * math.cos(math.radians(mid_lat))

            width_m  = (max_lon - min_lon) * m_per_deg_lon
            height_m = (max_lat - min_lat) * m_per_deg_lat
            area   = width_m * height_m

            print('Approximate area of query: ', area, ' sqm')

        elif t == MapLocation.POINT.name:
            if not (isinstance(v, tuple) and len(v) == 2):
                raise ValueError('POINT input must be [lat, lon]')
            
        elif t in [MapLocation.ADDRESS.name, MapLocation.PLACE.name]:
            if not isinstance(v, str):
                raise ValueError(f'{t} input must be a string')
        return v
    

    def query_key(self) -> str:
        '''
        Build a key for query parameters
        '''
        payload: dict[str, Any] = {
            'input_type':   self.input_type,
            'input':        self.input,
            'dist':         self.dist,
            
            # post-query config
            'lane_width':   self.lane_width,
        }

        js = json.dumps(payload, separators=(',', ':'), sort_keys=True)
        checksum = zlib.crc32(js.encode('utf-8')) & 0xFFFFFFFF
        return f'{checksum:08x}'


    # Settings: analysis specifications
    junc_trim_dist: float = Field(10.0, description='Distance from the center of the junction to be classified as junction')
    junc_ty_angle_threshold: float = Field(110.0, ge=0.0, description='Relative angle degree threshold for T- (angle > threshold) and Y- (angle <= threshold) junction classification')
    junc_right_hand_traffic: bool = Field(True, description='Whether vehicles drive on the right side of the road')
    junc_nbr_pos_threshold: float = Field(30.0, ge=0.0, description='Relative angle degree threshold determining neighbor edge position; Opposite (angle > 180-threshold), Right-angle (abs(angle - 90) < threshold)')
    junc_conflict_classifier: list[dict] = Field([
        { "this_move": "THRU",   "other_move": "THRU",   "nbr_pos": "OPP",  "conflict": "NO_CONFLICT" },
        { "this_move": "THRU",   "other_move": "THRU",   "nbr_pos": "NEAR", "conflict": "INTERSECT" },
        { "this_move": "THRU",   "other_move": "THRU",   "nbr_pos": "FAR",  "conflict": "INTERSECT" },

        { "this_move": "THRU",   "other_move": "TURN",   "nbr_pos": "OPP",  "conflict": "NO_CONFLICT" },
        { "this_move": "THRU",   "other_move": "TURN",   "nbr_pos": "NEAR", "conflict": "MERGE" },
        { "this_move": "THRU",   "other_move": "TURN",   "nbr_pos": "FAR",  "conflict": "NO_CONFLICT" },

        { "this_move": "THRU",   "other_move": "CROSS",  "nbr_pos": "OPP",  "conflict": "INTERSECT" },
        { "this_move": "THRU",   "other_move": "CROSS",  "nbr_pos": "NEAR", "conflict": "INTERSECT" },
        { "this_move": "THRU",   "other_move": "CROSS",  "nbr_pos": "FAR",  "conflict": "MERGE" },

        { "this_move": "THRU",   "other_move": "REVERSE","nbr_pos": "OPP",  "conflict": "MERGE" },
        { "this_move": "THRU",   "other_move": "REVERSE","nbr_pos": "NEAR", "conflict": "INTERSECT" },
        { "this_move": "THRU",   "other_move": "REVERSE","nbr_pos": "FAR",  "conflict": "INTERSECT" },

        { "this_move": "TURN",   "other_move": "THRU",   "nbr_pos": "OPP",  "conflict": "NO_CONFLICT" },
        { "this_move": "TURN",   "other_move": "THRU",   "nbr_pos": "NEAR", "conflict": "NO_CONFLICT" },
        { "this_move": "TURN",   "other_move": "THRU",   "nbr_pos": "FAR",  "conflict": "MERGE" },

        { "this_move": "TURN",   "other_move": "TURN",   "nbr_pos": "OPP",  "conflict": "NO_CONFLICT" },
        { "this_move": "TURN",   "other_move": "TURN",   "nbr_pos": "NEAR", "conflict": "NO_CONFLICT" },
        { "this_move": "TURN",   "other_move": "TURN",   "nbr_pos": "FAR",  "conflict": "NO_CONFLICT" },

        { "this_move": "TURN",   "other_move": "CROSS",  "nbr_pos": "OPP",  "conflict": "MERGE" },
        { "this_move": "TURN",   "other_move": "CROSS",  "nbr_pos": "NEAR", "conflict": "NO_CONFLICT" },
        { "this_move": "TURN",   "other_move": "CROSS",  "nbr_pos": "FAR",  "conflict": "NO_CONFLICT" },

        { "this_move": "TURN",   "other_move": "REVERSE","nbr_pos": "OPP",  "conflict": "NO_CONFLICT" },
        { "this_move": "TURN",   "other_move": "REVERSE","nbr_pos": "NEAR", "conflict": "MERGE" },
        { "this_move": "TURN",   "other_move": "REVERSE","nbr_pos": "FAR",  "conflict": "NO_CONFLICT" },

        { "this_move": "CROSS",  "other_move": "THRU",   "nbr_pos": "OPP",  "conflict": "INTERSECT" },
        { "this_move": "CROSS",  "other_move": "THRU",   "nbr_pos": "NEAR", "conflict": "MERGE" },
        { "this_move": "CROSS",  "other_move": "THRU",   "nbr_pos": "FAR",  "conflict": "INTERSECT" },

        { "this_move": "CROSS",  "other_move": "TURN",   "nbr_pos": "OPP",  "conflict": "MERGE" },
        { "this_move": "CROSS",  "other_move": "TURN",   "nbr_pos": "NEAR", "conflict": "NO_CONFLICT" },
        { "this_move": "CROSS",  "other_move": "TURN",   "nbr_pos": "FAR",  "conflict": "NO_CONFLICT" },

        { "this_move": "CROSS",  "other_move": "CROSS",  "nbr_pos": "OPP",  "conflict": "INTERSECT" },
        { "this_move": "CROSS",  "other_move": "CROSS",  "nbr_pos": "NEAR", "conflict": "INTERSECT" },
        { "this_move": "CROSS",  "other_move": "CROSS",  "nbr_pos": "FAR",  "conflict": "INTERSECT" },

        { "this_move": "CROSS",  "other_move": "REVERSE","nbr_pos": "OPP",  "conflict": "INTERSECT" },
        { "this_move": "CROSS",  "other_move": "REVERSE","nbr_pos": "NEAR", "conflict": "INTERSECT" },
        { "this_move": "CROSS",  "other_move": "REVERSE","nbr_pos": "FAR",  "conflict": "MERGE" },

        { "this_move": "REVERSE","other_move": "THRU",   "nbr_pos": "OPP",  "conflict": "MERGE" },
        { "this_move": "REVERSE","other_move": "THRU",   "nbr_pos": "NEAR", "conflict": "INTERSECT" },
        { "this_move": "REVERSE","other_move": "THRU",   "nbr_pos": "FAR",  "conflict": "INTERSECT" },

        { "this_move": "REVERSE","other_move": "TURN",   "nbr_pos": "OPP",  "conflict": "NO_CONFLICT" },
        { "this_move": "REVERSE","other_move": "TURN",   "nbr_pos": "NEAR", "conflict": "NO_CONFLICT" },
        { "this_move": "REVERSE","other_move": "TURN",   "nbr_pos": "FAR",  "conflict": "MERGE" },

        { "this_move": "REVERSE","other_move": "CROSS",  "nbr_pos": "OPP",  "conflict": "INTERSECT" },
        { "this_move": "REVERSE","other_move": "CROSS",  "nbr_pos": "NEAR", "conflict": "MERGE" },
        { "this_move": "REVERSE","other_move": "CROSS",  "nbr_pos": "FAR",  "conflict": "INTERSECT" },

        { "this_move": "REVERSE","other_move": "REVERSE","nbr_pos": "OPP","conflict":"INTERSECT" },
        { "this_move": "REVERSE","other_move": "REVERSE","nbr_pos": "NEAR","conflict":"INTERSECT" },
        { "this_move": "REVERSE","other_move": "REVERSE","nbr_pos": "FAR","conflict":"INTERSECT" }
        ], description='Conflict classification based on current vehicle movement and position and other vehicle position and movement on another neighbor edge')
    
    # Settings: 
    lane_width: float = Field(4.0, description='Default lane width of the edge')

    @model_validator(mode='before')
    def apply_default_settings(cls, data: dict[str,Any]) -> dict[str,Any]:
        if data.get('default_settings'):
            return data

        wb    = WebRequest._EXCEL_WB()
        sheet = wb[cls.SHEET_SETTINGS]

        # build a list of rows as dicts
        rows = []
        for feat_coll, feat_type, attr, val in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({
                'collection': feat_coll,
                'feature':    feat_type,
                'attr':       attr,
                'value':      val
            })

        for row in rows:
            a = row['attr']
            v = row['value']

            if a == 'junc_trim_dist':
                data['junc_trim_dist'] = float(v)
            elif a == 'junc_ty_angle_threshold':
                data['junc_ty_angle_threshold'] = float(v)
            elif a == 'junc_right_hand_traffic':
                data['junc_right_hand_traffic'] = bool(v)
            elif a == 'junc_nbr_pos_threshold':
                data['junc_nbr_pos_threshold'] = float(v)
            elif a == 'lane_width':
                data['lane_width'] = float(v)
            # if you had classifier in Excel, you could parse JSON here too

        return data
    
    @model_validator(mode='before')
    def apply_default_conflicts(cls, data: dict[str,Any]) -> dict[str,Any]:
        if not data.get('default_conflict_classifier'):
            wb    = WebRequest._EXCEL_WB()
            sheet = wb[cls.SHEET_CONFLICT]

            # Build list of dicts from the four columns
            rules: List[dict[str,str]] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                this_move, other_move, nbr_pos, conflict = row
                # skip any incomplete rows
                if None in (this_move, other_move, nbr_pos, conflict):
                    continue
                rules.append({
                    'this_move': str(this_move).strip(),
                    'other_move': str(other_move).strip(),
                    'nbr_pos': str(nbr_pos).strip(),
                    'conflict': str(conflict).strip(),
                })

            data['junc_conflict_classifier'] = rules

        return data


    # ODD
    odd: dict[str, list] = Field({
        'junction_type':        ['ALL'],
        'junction_conflict':    ['ALL'],
    }, description='Operational Design Domain criteria')


    @model_validator(mode='before')
    def apply_default_odd(cls, data: dict[str,Any]) -> dict[str,Any]:
        if not data.get('odd_all'):
            wb    = WebRequest._EXCEL_WB()
            sheet = wb[cls.SHEET_ODD]

            odd_map: dict[str,list] = {}
            for feat_coll, feat_type, attr, val in sheet.iter_rows(min_row=2, values_only=True):
                if not isinstance(feat_type, str):
                    continue
                odd_map.setdefault(str(attr), []).append(str(val))

            if odd_map:
                data['odd'] = odd_map

        return data
    

    def req_key(self) -> str:
        '''
        Build a key for query parameters
        '''
        payload = self.model_dump()
        js = json.dumps(payload, separators=(',', ':'), sort_keys=True)
        checksum = zlib.crc32(js.encode('utf-8')) & 0xFFFFFFFF
        return f'{checksum:08x}'

    
def parse_geometry(v: Any) -> BaseGeometry:
    '''
    Accepts:
      a Shapely geometry
      a GeoJSON dict {"type":..., "coordinates":...}
      a WKT string
    Returns a Shapely geometry.
    '''
    if isinstance(v, BaseGeometry):
        return v
    if isinstance(v, dict):
        return shape(v)
    if isinstance(v, str):
        return wkt.loads(v)
    raise TypeError(f'Can\'t parse geometry for output')
    


class BaseExporter(BaseModel):
    '''
    Single feature export base model
    GeoJson compliant format
    '''
    type: str = Field(..., description='Type of data')
    geometry: BaseGeometry = Field(..., description='Geometry object')
    properties: dict[str, Any] = Field(..., description='Metadata')


    @field_validator('geometry', mode='before')
    def _load_geometry(cls, v):
        return parse_geometry(v)
    

    @field_validator('properties')
    def _structure_properties(cls, v):
        if not isinstance(v, dict):
            raise ValueError('`properties` must be a dictionary')
        feature_type = v.get('feature_type')
        if not isinstance(feature_type, str) or not feature_type:
            raise ValueError('`properties` must include a non-empty string specifying `feature_type`')
        metadata = v.get('metadata')
        if not metadata:
            metadata = {k: val for k, val in v.items() if k != 'feature_type'}
        return{
            'feature_type': feature_type,
            'metadata': metadata,
        }


    model_config = ConfigDict(
        extra='allow',
        arbitrary_types_allowed=True,
        json_encoders = {
            BaseGeometry: lambda g: {
                'type':         g.geom_type,
                'coordinates':  mapping(g)['coordinates']
            }
        },
    )


class Feature:
    '''
    Model for constructing the list of elements related to a single feature attribute
    Used for exporting feature dictionary to frontend
    '''

    def __init__(self, feature_attr: str, values: List):
        self.feature_attr = feature_attr
        self.values = values

    @staticmethod
    def enum_values(enum_cls: type[Enum]) -> list[str]:
        values = []
        for e in enum_cls:
            val = e.value
            # Use value if it's a str and not equal to the name
            if isinstance(val, str) and val != e.name:
                values.append(val)
            else:
                values.append(e.name)
        return values


    def as_dict(self) -> dict:
        return {'feature_attr': self.feature_attr, 'values': self.values}


class FeatureDict:
    '''
    Model for aggregating all feature attributes and related elements by feature type
    Used for exporting feature dictionary to frontend
    '''
    def __init__(self, base_map: List[dict] = []):
        self.feature_map = base_map if base_map else []

    def add_feature_type(self, feature_type: str, features: List):
        self.feature_map.append({
            'feature_type': feature_type,
            'features': [f.as_dict() for f in features]
        })

    def out(self) -> List[dict]:
        return self.feature_map
    

# Sample output
# {
#     'feature_type': 'junction',
#     'features': [
#         {'feature_attr': 'junction_type',
#          'values': ['T_JUNCTION', 'ROUNDABOUT']},
#         {'feature_attr': 'junction_conflict',
#          'values': ['INTERSECT', 'MERGE', 'NO_CONFLICT']}
#     ]
# }


# Sample usage
# junction_features = [
#     Feature('junction_type', ['T_JUNCTION', 'CROSS_JUNCTION']),
#     Feature('junction_conflict', ['NO_CONFLICT', 'PEDESTRIAN_CONFLICT'])
# ]

# lanes_features = [
#     Feature('max_speed', []),
#     Feature('highway_type', ['motor_way', 'trunk_way'])
# ]

# fd = FeatureDict2()
# fd.add_feature_type('junction', junction_features)
# fd.add_feature_type('lanes', lanes_features)
